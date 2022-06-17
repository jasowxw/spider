# -*- coding: UTF-8 -*-
'''
@Project ：spider_demo 
@File    ：demo.py
@IDE     ：PyCharm 
@Author  ：闻小文
@Date    ：2021/10/16 1:35 
@win_name   ：wzw
'''
import time
from concurrent.futures import ThreadPoolExecutor
from threading import Lock

import requests
import openpyxl

lock = Lock()  # 生成一个锁


def post_():
    """一个生成器，每一次调用返回一页数据的post请求"""
    url = 'http://www.xinfadi.com.cn/getPriceData.html'  # 通过post请求改变表单的数据，URL并不变化
    for i in range(1, 100):
        data = {'limit': '20', 'current': i}
        yield requests.post(url=url, data=data)


def dwon(response):
    list = response.json()['list']
    row = []  # 用以存放一页数据
    for i in list:
        row_ = [i['prodCat'], i['prodPcat'], i['prodName'], i['lowPrice'], i['highPrice'], i['avgPrice'], i['specInfo'],
                i['place'], i['unitInfo'], i['pubDate']]
        row.append(row_)
    print(row)
    with lock:
        wb = openpyxl.load_workbook('price.xlsx')
        for i in row:
            wb.active.append(i)
        wb.save('price.xlsx')




def get_time(fu):
    def war():
        start = time.time()
        fu()
        end = time.time()
        print(f'时间为：{end-start}')

    return war


@get_time
def main():
    title = ['一级分类', '二级分类', '品名','最低价', '最高价', '平均价', '规格', '产地', '单位', '发布日期']
    book = openpyxl.Workbook()
    for i in range(len(title)):
        book.active.cell(1, i + 1, title[i])
    book.save('price.xlsx')
    with ThreadPoolExecutor(48) as f:
        for response in post_():
            f.submit(dwon,response)
if __name__ == '__main__':
    main()



